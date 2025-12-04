from fastapi import FastAPI, HTTPException, File, UploadFile
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import os
from dotenv import load_dotenv
import json
import io
import csv
from typing import Optional, List

# Import database and file storage
try:
    from models.database import (
        init_db, create_chat_session, add_chat_message, get_chat_messages,
        save_file_upload, save_classification_results, get_file_uploads,
        get_all_chat_sessions, delete_chat_session, update_chat_session_title,
        get_statistics
    )
    from utils.file_storage import save_uploaded_file, get_file_size, init_upload_directory
    DATABASE_ENABLED = True
except ImportError as e:
    print(f"Warning: Database not available: {e}")
    DATABASE_ENABLED = False

# Try to import openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None

# Load env vars (reuses the same environment variables as bug_classifier)
load_dotenv()

# Import classifier (should be import-safe now)
try:
    # batch_classify is the new async batch classifier; keep classify_bug for single-item fallback
    from services.classifier import batch_classify, classify_bug
except ImportError:
    # fallback to relative import if running as package
    try:
        from .services.classifier import batch_classify, classify_bug
    except ImportError as e:
        raise RuntimeError(f"Failed to import classifier: {e}")

# Try to import pandas and openpyxl for Excel support
try:
    import pandas as pd
except ImportError:
    pd = None

# Import Jira fetcher
try:
    from fetch_jira import fetch_jira_issues
except ImportError:
    try:
        from .fetch_jira import fetch_jira_issues
    except ImportError as e:
        # This is not a critical error, so we can just print a warning
        print(f"Warning: Jira fetcher not available: {e}")
        fetch_jira_issues = None

app = FastAPI(title="BugClassifier API", version="0.1")

# Initialize database and upload directory on startup
@app.on_event("startup")
async def startup_event():
    if DATABASE_ENABLED:
        init_db()
        init_upload_directory()
        print("Database and file storage initialized")

# Store latest classification results in memory for download (session-based)
latest_results = None

# Allow the Web frontend (served from localhost:5173) to call this API during development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173", "http://localhost:5174/"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

class ClassifyRequest(BaseModel):
    text: str
    model: str = "GPT-5"  # Default to GPT-5


class ClassifyItem(BaseModel):
    text: str
    label: str
    raw: str
    team: str | None = None
    severity: str | None = None


class ClassifyResponse(BaseModel):
    results: list[ClassifyItem]


class UploadResponse(BaseModel):
    filename: str
    total_rows: int
    classified_rows: int
    results: list[ClassifyItem]
    file_upload_id: Optional[int] = None


class DownloadExcelRequest(BaseModel):
    results: list[ClassifyItem]

class CreateSessionRequest(BaseModel):
    session_id: str
    title: str = "Untitled"

class UpdateSessionRequest(BaseModel):
    title: str

class AddMessageRequest(BaseModel):
    role: str
    content: str
    file_upload_id: Optional[int] = None
    model: Optional[str] = None

class JiraRequest(BaseModel):
    jql: str

@app.get("/health")
async def health():
    return {"status": "ok"}

@app.post("/classify", response_model=ClassifyResponse)
async def classify(req: ClassifyRequest):
    if not req.text or not req.text.strip():
        raise HTTPException(status_code=400, detail="text is required")

    # Split input into non-empty lines; treat each line as a separate bug report
    lines = [l.strip() for l in req.text.splitlines() if l.strip()]
    if not lines:
        raise HTTPException(status_code=400, detail="no bug lines found in text")

    # Validate model
    model = req.model if req.model in ["GPT-5", "Llama"] else "GPT-5"

    try:
        # Use classify_bug for single line, batch_classify for multiple
        if len(lines) == 1:
            result = await classify_bug(lines[0], model=model)
            classified = [result]
        else:
            classified = await batch_classify(lines, model=model)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"classification error: {e}")

    results = []
    for i, l in enumerate(lines):
        res = classified[i] if i < len(classified) else None
        if isinstance(res, dict):
            label = res.get('label') or ''
            raw = res.get('reason') or json.dumps(res)
            team = res.get('team')
            severity = res.get('severity')
        else:
            label = res or ''
            raw = json.dumps(res)
            team = None
            severity = None
        results.append({"text": l, "label": label, "raw": raw, "team": team, "severity": severity})

    return {"results": results}


@app.post("/upload", response_model=UploadResponse)
async def upload_file(file: UploadFile = File(...), session_id: Optional[str] = None, model: str = "GPT-5"):
    """
    Upload CSV/XLSX file với columns: No, Nội dung bug
    Trả về kết quả phân loại: No, Nội dung bug, Label, Reason, Team, Severity
    Optional: session_id để liên kết với chat session
    Optional: model - "GPT-5" hoặc "Llama" (default: "GPT-5")
    """
    global latest_results
    
    if not file.filename:
        raise HTTPException(status_code=400, detail="no file provided")
    
    try:
        content = await file.read()
        # Save file to disk if database is enabled
        file_path = None
        file_size = len(content)
        if DATABASE_ENABLED:
            file_path = save_uploaded_file(content, file.filename)
            file_size = get_file_size(file_path)
        
        # Parse CSV or Excel
        bugs_with_rows = []
        
        if file.filename.endswith('.csv'):
            stream = io.StringIO(content.decode('utf-8'))
            reader = csv.DictReader(stream)
            rows = list(reader)
            
            # Nối toàn bộ cột thành 1 chuỗi
            for row in rows:
                # Clean row: thay thế nan, None, empty bằng empty string
                cleaned_row = {k: ('' if v is None or str(v).lower() == 'nan' or not str(v).strip() else str(v).strip()) for k, v in row.items()}
                # Nối tất cả giá trị cột thành 1 chuỗi, phân cách bằng | 
                combined_text = " | ".join([f"{k}: {v}" for k, v in cleaned_row.items() if v])
                bugs_with_rows.append({
                    'original_row': cleaned_row,
                    'combined_text': combined_text
                })
        
        elif file.filename.endswith(('.xlsx', '.xls')):
            if pd is None:
                raise HTTPException(status_code=400, detail="pandas not installed for Excel support")
            
            df = pd.read_excel(io.BytesIO(content))
            
            # Nối toàn bộ cột thành 1 chuỗi
            for idx, row in df.iterrows():
                # Clean row: thay thế NaN, None bằng empty string
                cleaned_row = {k: ('' if pd.isna(v) else str(v).strip()) for k, v in row.items()}
                # Nối tất cả giá trị cột thành 1 chuỗi
                combined_text = " | ".join([f"{k}: {v}" for k, v in cleaned_row.items() if v])
                bugs_with_rows.append({
                    'original_row': cleaned_row,
                    'combined_text': combined_text
                })
        
        else:
            raise HTTPException(status_code=400, detail="only CSV and Excel files are supported")
        
        if not bugs_with_rows:
            raise HTTPException(status_code=400, detail="no valid bug entries found in file")
        
        # Extract combined texts for classification
        bug_texts = [item['combined_text'] for item in bugs_with_rows]
        
        # Validate model parameter
        if model not in ["GPT-5", "Llama"]:
            raise HTTPException(status_code=400, detail=f"Invalid model: {model}. Must be 'GPT-5' or 'Llama'")
        
        # Classify using batch with selected model
        classified = await batch_classify(bug_texts, model=model)
        
        # Build results
        results = []
        for i, (bug_entry, classification) in enumerate(zip(bugs_with_rows, classified)):
            # Handle None classification
            if classification is None:
                classification = {"label": "", "reason": "classification_failed", "team": None, "severity": None}
            
            if isinstance(classification, dict):
                label = classification.get('label') or ''
                reason = classification.get('reason') or ''
                team = classification.get('team')
                severity = classification.get('severity')
            else:
                label = str(classification) or ''
                reason = ''
                team = None
                severity = None
            
            results.append({
                "text": bug_entry['combined_text'],
                "label": label,
                "raw": reason,
                "team": team,
                "severity": severity
            })
        
        latest_results = {
            "filename": file.filename,
            "results": results,
            "bugs_with_no": [item['original_row'] for item in bugs_with_rows],
            "classifications": classified
        }
        # Save to database if enabled (optional)
        upload_id = None
        if DATABASE_ENABLED and file_path:
            try:
                upload_id = save_file_upload(
                    session_id=session_id,
                    filename=file.filename,
                    file_path=file_path,
                    file_size=file_size,
                    total_rows=len(bugs_with_rows),
                    classified_rows=len(results)
                )
                # Truyền original_rows để lưu toàn bộ dữ liệu gốc
                original_rows = [item['original_row'] for item in bugs_with_rows]
                save_classification_results(upload_id, results, original_rows)
            except Exception as e:
                print(f"Warning: Failed to save to database: {e}")
        return {
            "filename": file.filename,
            "total_rows": len(bugs_with_rows),
            "classified_rows": len(results),
            "results": results,
            "file_upload_id": upload_id
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"file processing error: {str(e)}")


@app.get("/download-result")
async def download_result():
    """
    Download kết quả phân loại dưới dạng CSV
    Columns: No, Nội dung bug, Label, Memo, Team, Severity
    """
    global latest_results
    
    if not latest_results:
        raise HTTPException(status_code=400, detail="no classification results available")
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Lấy các cột gốc từ row đầu tiên
    original_rows = latest_results['bugs_with_no']
    if not original_rows:
        raise HTTPException(status_code=400, detail="no data available")
    
    # Lấy tên các cột gốc
    original_columns = list(original_rows[0].keys()) if isinstance(original_rows[0], dict) else []
    
    # Write header: các cột gốc + các cột bổ sung
    header = original_columns + ['Label', 'Memo', 'Team', 'Severity']
    writer.writerow(header)
    
    # Write data rows
    for i, (bug_entry, classification) in enumerate(zip(original_rows, latest_results['classifications'])):
        if isinstance(classification, dict):
            label = classification.get('label') or ''
            reason = classification.get('reason') or ''
            team = classification.get('team') or ''
            severity = classification.get('severity') or ''
        else:
            label = str(classification) or ''
            reason = ''
            team = ''
            severity = ''
        
        # Kết hợp dữ liệu gốc và kết quả phân loại
        row_data = []
        if isinstance(bug_entry, dict):
            for col in original_columns:
                val = bug_entry.get(col, '')
                # Thay thế nan bằng empty string
                val_str = str(val) if val is not None else ''
                row_data.append('' if val_str.lower() == 'nan' else val_str)
        row_data.extend([label, reason, team, severity])
        
        writer.writerow(row_data)
    
    # Return as file download
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=classification_result.csv"}
    )


@app.post("/download-excel")
async def download_excel(req: DownloadExcelRequest):
    """
    Generate Excel file từ classification results
    """
    global latest_results
    
    if not Workbook:
        raise HTTPException(status_code=400, detail="openpyxl not installed")
    
    if not req.results or not latest_results:
        raise HTTPException(status_code=400, detail="no results to download")
    
    # Lấy các cột gốc từ file upload
    original_rows = latest_results['bugs_with_no']
    if not original_rows:
        raise HTTPException(status_code=400, detail="no data available")
    
    # Lấy tên các cột gốc
    original_columns = list(original_rows[0].keys()) if isinstance(original_rows[0], dict) else []
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Classification Results"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    # Màu cam cho các cột bổ sung
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    orange_header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    
    # Write headers: các cột gốc + các cột bổ sung
    all_headers = original_columns + ['Label', 'Memo', 'Team', 'Severity']
    num_original_cols = len(original_columns)
    
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        # Nếu là cột bổ sung thì tô màu cam
        if col_idx > num_original_cols:
            cell.fill = orange_header_fill
        else:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Sắp xếp kết quả theo label
    sorted_results = sorted(req.results, key=lambda r: (r.label or '').lower())
    
    # Write data rows
    for row_idx, (result, original_row) in enumerate(zip(sorted_results, original_rows), 2):
        col_idx = 1
        
        # Ghi các cột gốc
        if isinstance(original_row, dict):
            for col_name in original_columns:
                cell = ws.cell(row=row_idx, column=col_idx)
                val = original_row.get(col_name, '')
                # Thay thế nan bằng empty string
                val_str = str(val) if val is not None else ''
                cell.value = '' if val_str.lower() == 'nan' else val_str
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                col_idx += 1
        
        # Ghi các cột bổ sung (không tô màu cam cho data)
        added_data = [
            result.label,
            result.raw,
            result.team or '',
            result.severity or ''
        ]
        for value in added_data:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            col_idx += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, len(all_headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        if col_idx <= num_original_cols:
            ws.column_dimensions[col_letter].width = 15
        else:
            # Các cột bổ sung rộng hơn
            if col_idx == num_original_cols + 1:  # Label
                ws.column_dimensions[col_letter].width = 15
            elif col_idx == num_original_cols + 2:  # Memo
                ws.column_dimensions[col_letter].width = 35
            elif col_idx == num_original_cols + 3:  # Team
                ws.column_dimensions[col_letter].width = 20
            else:  # Severity
                ws.column_dimensions[col_letter].width = 12
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return StreamingResponse(
        iter([excel_buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=classification_results.xlsx"}
    )


# ===== Chat Session Management Endpoints =====

@app.post("/chat/sessions")
async def create_session(request: CreateSessionRequest):
    """Create a new chat session"""
    if not DATABASE_ENABLED:
        raise HTTPException(status_code=503, detail="database not available")
    
    success = create_chat_session(request.session_id, request.title)
    if not success:
        raise HTTPException(status_code=409, detail="session already exists")
    
    return {"session_id": request.session_id, "title": request.title, "success": True}


@app.get("/chat/sessions")
async def list_sessions():
    """Get all chat sessions"""
    if not DATABASE_ENABLED:
        raise HTTPException(status_code=503, detail="database not available")
    
    sessions = get_all_chat_sessions()
    return {"sessions": sessions}


@app.put("/chat/sessions/{session_id}")
async def update_session(session_id: str, request: UpdateSessionRequest):
    """Update chat session title"""
    if not DATABASE_ENABLED:
        raise HTTPException(status_code=503, detail="database not available")
    
    success = update_chat_session_title(session_id, request.title)
    if not success:
        raise HTTPException(status_code=404, detail="session not found")
    
    return {"session_id": session_id, "title": request.title, "success": True}


@app.delete("/chat/sessions/{session_id}")
async def delete_session(session_id: str):
    """Delete a chat session and all related data"""
    if not DATABASE_ENABLED:
        raise HTTPException(status_code=503, detail="database not available")
    
    success = delete_chat_session(session_id)
    if not success:
        raise HTTPException(status_code=404, detail="session not found")
    
    return {"deleted": True}


@app.post("/chat/sessions/{session_id}/messages")
async def add_message(session_id: str, request: AddMessageRequest):
    """Add a message to chat session"""
    if not DATABASE_ENABLED:
        raise HTTPException(status_code=503, detail="database not available")
    
    if request.role not in ["user", "assistant"]:
        raise HTTPException(status_code=400, detail="role must be 'user' or 'assistant'")
    
    message_id = add_chat_message(session_id, request.role, request.content, request.file_upload_id, request.model)
    return {"message_id": message_id, "session_id": session_id, "success": True}


@app.get("/chat/sessions/{session_id}/messages")
async def get_messages(session_id: str, limit: Optional[int] = None):
    """Get all messages for a chat session"""
    if not DATABASE_ENABLED:
        raise HTTPException(status_code=503, detail="database not available")
    
    messages = get_chat_messages(session_id, limit)
    return {"messages": messages}


@app.get("/uploads")
async def list_uploads(session_id: Optional[str] = None, limit: int = 50):
    """Get file uploads, optionally filtered by session"""
    if not DATABASE_ENABLED:
        raise HTTPException(status_code=503, detail="database not available")
    
    uploads = get_file_uploads(session_id, limit)
    return {"uploads": uploads}


@app.get("/statistics")
async def get_stats():
    """Get statistics including vector store stats"""
    if not DATABASE_ENABLED:
        raise HTTPException(status_code=503, detail="database not available")
    
    stats = get_statistics()
    
    # Add vector store stats if available
    try:
        from models.vector_store import get_vector_store_stats
        vector_stats = get_vector_store_stats()
        stats['vector_store'] = vector_stats
    except:
        stats['vector_store'] = {"available": False}
    
    return stats


@app.get("/classification-results/{file_upload_id}")
async def get_classification_by_upload_id(file_upload_id: int):
    """Get classification results for a file upload"""
    if not DATABASE_ENABLED:
        raise HTTPException(status_code=503, detail="database not available")
    
    try:
        from models.database import get_classification_results
        db_results = get_classification_results(file_upload_id)
        
        # Map DB results to API model
        results = []
        for r in db_results:
            results.append({
                "text": r.get('bug_text', ''),
                "label": r.get('label', ''),
                "raw": r.get('reason', ''),
                "team": r.get('team'),
                "severity": r.get('severity')
            })
            
        return {"file_upload_id": file_upload_id, "results": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/jira/fetch")
async def fetch_jira(req: JiraRequest):
    """
    Fetch issues from Jira based on a JQL query.
    """
    if not fetch_jira_issues:
        raise HTTPException(status_code=501, detail="Jira integration is not configured on the server.")

    if not req.jql or not req.jql.strip():
        raise HTTPException(status_code=400, detail="JQL query is required.")

    try:
        issues = fetch_jira_issues(req.jql)
        if "error" in issues:
            raise HTTPException(status_code=500, detail=issues.get("details", issues["error"]))
        return {"issues": issues}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")


# ============================================================================
# ChromaDB Vector Store Endpoints
# ============================================================================

@app.get("/vector/stats")
async def vector_stats():
    """Get ChromaDB vector store statistics"""
    try:
        from models.vector_store import get_vector_store_stats
        stats = get_vector_store_stats()
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get stats: {str(e)}")


# uvicorn entrypoint hint
# run with: uvicorn Server.api:app --reload --port 8000

